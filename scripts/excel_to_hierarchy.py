"""
Read comprehensive plan Excel export and emit nested JSON hierarchy.
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

# Expected header labels (sheet may use typo "Sub Policy Decription")
HEADER_ALIASES = {
    "chapter number": "chapterNumber",
    "chapter title": "chapterTitle",
    "goal number": "goalNumber",
    "goal description": "goalDescription",
    "goal detail": "goalDetail",
    "policy number": "policyNumber",
    "policy description": "policyDescription",
    "sub policy": "subPolicy",
    "sub policy decription": "subPolicyDescription",
    "sub policy description": "subPolicyDescription",
    "sub policy text": "subPolicyText",
    "sub policy sub level number": "subLevelRoman",
    "sub policy sub level description": "subLevelDescription",
}


def norm_scalar(v):
    if v is None:
        return None
    if isinstance(v, float) and v == int(v):
        v = int(v)
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    return v


def norm_key_str(v) -> str | None:
    n = norm_scalar(v)
    if n is None:
        return None
    return str(n)


def load_header_map(ws) -> dict[str, int]:
    row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_by_field: dict[str, int] = {}
    for idx, cell in enumerate(row1):
        if cell is None:
            continue
        label = str(cell).strip().lower()
        field = HEADER_ALIASES.get(label)
        if field:
            col_by_field[field] = idx
    required = [
        "chapterNumber",
        "chapterTitle",
        "goalNumber",
        "goalDescription",
        "goalDetail",
        "policyNumber",
        "policyDescription",
        "subPolicy",
        "subPolicyDescription",
        "subPolicyText",
        "subLevelRoman",
        "subLevelDescription",
    ]
    missing = [f for f in required if f not in col_by_field]
    if missing:
        raise SystemExit(f"Missing columns in sheet: {missing}. Found: {sorted(col_by_field)}")
    return col_by_field


def row_dict(row: tuple, col_by_field: dict[str, int]) -> dict:
    out = {}
    for field, idx in col_by_field.items():
        out[field] = row[idx] if idx < len(row) else None
    return out


def build_hierarchy(rows: list[dict]) -> tuple[list[dict], dict]:
    chapters: dict = {}
    chapter_order: list = []

    for r in rows:
        ch_num = norm_scalar(r["chapterNumber"])
        ch_title = norm_key_str(r["chapterTitle"])
        g_num = norm_key_str(r["goalNumber"])
        g_desc = norm_key_str(r["goalDescription"])
        g_detail = norm_key_str(r["goalDetail"])
        p_num = norm_key_str(r["policyNumber"])
        p_desc = norm_key_str(r["policyDescription"])

        sp_letter = norm_key_str(r["subPolicy"])
        sp_desc = norm_key_str(r["subPolicyDescription"])
        sp_text = norm_key_str(r["subPolicyText"])
        sl_roman = norm_key_str(r["subLevelRoman"])
        sl_desc = norm_key_str(r["subLevelDescription"])

        if ch_num is None and ch_title is None:
            continue

        ck = (ch_num, ch_title)
        if ck not in chapters:
            chapters[ck] = {
                "chapterNumber": ch_num,
                "chapterTitle": ch_title,
                "_goal_order": [],
                "goals": {},
            }
            chapter_order.append(ck)

        ch = chapters[ck]
        gk = (g_num, g_desc)
        if gk not in ch["goals"]:
            ch["goals"][gk] = {
                "goalNumber": g_num,
                "goalDescription": g_desc,
                "_detail_order": [],
                "details": {},
            }
            ch["_goal_order"].append(gk)

        goal = ch["goals"][gk]
        dk = g_detail if g_detail is not None else ""
        if dk not in goal["details"]:
            goal["details"][dk] = {
                "detail": g_detail,
                "_policy_order": [],
                "policies": {},
            }
            goal["_detail_order"].append(dk)

        det = goal["details"][dk]
        if p_num is None and p_desc is None:
            continue

        pk = (p_num, p_desc)
        if pk not in det["policies"]:
            det["policies"][pk] = {
                "policyNumber": p_num,
                "policyDescription": p_desc,
                "_sub_order": [],
                "_subs": {},
            }
            det["_policy_order"].append(pk)

        pol = det["policies"][pk]

        sub_key = (sp_letter, sp_desc, sp_text)
        if sub_key not in pol["_subs"]:
            pol["_subs"][sub_key] = {
                "letter": sp_letter,
                "description": sp_desc,
                "text": sp_text,
                "_sublevel_keys": [],
                "_sublevels": {},
            }
            pol["_sub_order"].append(sub_key)

        sub = pol["_subs"][sub_key]
        if sl_roman is not None or sl_desc is not None:
            slk = (sl_roman, sl_desc)
            if slk not in sub["_sublevels"]:
                sub["_sublevels"][slk] = {"roman": sl_roman, "description": sl_desc}
                sub["_sublevel_keys"].append(slk)

    def finalize_sub(sub: dict) -> dict:
        out = {}
        if sub["letter"] is not None:
            out["letter"] = sub["letter"]
        if sub["description"] is not None:
            out["description"] = sub["description"]
        if sub["text"] is not None:
            out["text"] = sub["text"]
        levels = [sub["_sublevels"][k] for k in sub["_sublevel_keys"]]
        if levels:
            out["subLevels"] = levels
        return out

    def finalize_policy(pol: dict) -> dict:
        base = {}
        if pol["policyNumber"] is not None:
            base["policyNumber"] = pol["policyNumber"]
        if pol["policyDescription"] is not None:
            base["policyDescription"] = pol["policyDescription"]
        subs = [finalize_sub(pol["_subs"][k]) for k in pol["_sub_order"]]
        if subs:
            base["subPolicies"] = subs
        return base

    def finalize_detail(det: dict) -> dict:
        item = {}
        if det["detail"] is not None:
            item["detail"] = det["detail"]
        policies = [finalize_policy(det["policies"][k]) for k in det["_policy_order"]]
        if policies:
            item["policies"] = policies
        return item

    def finalize_goal(goal: dict) -> dict:
        base = {}
        if goal["goalNumber"] is not None:
            base["goalNumber"] = goal["goalNumber"]
        if goal["goalDescription"] is not None:
            base["goalDescription"] = goal["goalDescription"]
        details = [finalize_detail(goal["details"][k]) for k in goal["_detail_order"]]
        base["goalDetails"] = details
        return base

    out_chapters = []
    stats = {"totalGoals": 0, "totalPolicies": 0}

    for ck in chapter_order:
        ch = chapters[ck]
        goals = [finalize_goal(ch["goals"][gk]) for gk in ch["_goal_order"]]
        stats["totalGoals"] += len(goals)
        for g in goals:
            for gd in g.get("goalDetails", []):
                stats["totalPolicies"] += len(gd.get("policies", []))
        oc = {
            "chapterNumber": ch["chapterNumber"],
            "chapterTitle": ch["chapterTitle"],
            "goals": goals,
        }
        out_chapters.append(oc)

    return out_chapters, stats


def main() -> None:
    root = Path(r"C:\Users\e25347\CABQ-Comprehensive-Plan-Action-App")
    xlsx = Path(r"C:\Users\e25347\Downloads\comprehensive plan table.xlsx")
    out_json = root / "public" / "data" / "comprehensive-plan-hierarchy.json"

    if not xlsx.is_file():
        print(f"ERROR: Excel file not found: {xlsx}", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb.active
    col_by_field = load_header_map(ws)

    data_rows: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(c is None or (isinstance(c, str) and not str(c).strip()) for c in row):
            continue
        data_rows.append(row_dict(tuple(row), col_by_field))
    wb.close()

    chapters, stats = build_hierarchy(data_rows)
    payload = {"chapters": chapters}

    out_json.parent.mkdir(parents=True, exist_ok=True)
    out_json.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    n_ch = len(chapters)
    print(f"Chapters: {n_ch}")
    print(f"Total goals: {stats['totalGoals']}")
    print(f"Total policies: {stats['totalPolicies']}")
    print(f"Wrote: {out_json}")


if __name__ == "__main__":
    main()
